import os
import uuid
import time
from io import BytesIO
from datetime import datetime, date
from typing import Optional, List, Dict, Any, Tuple

import requests
import openpyxl
from fastapi import FastAPI, UploadFile, File, Form, Header, HTTPException
from fastapi.responses import StreamingResponse

app = FastAPI()

ARCGIS_QUERY_URL = (
    "https://arcgis.gdsc.miami.edu/arcgis/rest/services/"
    "mdc_property_point_view/FeatureServer/0/query"
)

# Keep outFields explicit (outFields="*" can 400 on this layer)
DEFAULT_OUTFIELDS = (
    "objectid,folio,true_site_addr,true_site_unit,true_site_city,true_site_zip_code,"
    "true_owner1,true_owner2,true_owner3,"
    "true_mailing_addr1,true_mailing_addr2,true_mailing_addr3,"
    "true_mailing_city,true_mailing_state,true_mailing_zip_code,true_mailing_country,"
    "condo_flag,parent_folio"
)

# Condo tower plausibility threshold
MIN_TOWER_ROWS_FOR_COMPLETED = 50


def _require_api_key(x_api_key: Optional[str]) -> None:
    expected = os.environ.get("API_KEY")
    if not expected:
        raise HTTPException(status_code=500, detail="Server misconfigured: API_KEY not set")
    if not x_api_key or x_api_key != expected:
        raise HTTPException(status_code=401, detail="Invalid API key")


def _classify_owner_type(owner_raw: str) -> str:
    s = (owner_raw or "").upper()
    if any(k in s for k in ["LLC", "INC", "CORP", " CO", " LTD", "LP", "LLP", "HOLDINGS", "PARTNERS"]):
        return "Entity"
    if any(k in s for k in ["TRUST", "TRUSTEE"]):
        return "Trust"
    if s.strip():
        return "Individual"
    return "Unknown"


def _join_owner(*parts: Optional[str]) -> str:
    vals = [p.strip() for p in parts if isinstance(p, str) and p.strip()]
    return "|".join(vals)


def _arcgis_page(where: str, out_fields: str, offset: int, count: int) -> Dict[str, Any]:
    """
    Fetch one paginated page from ArcGIS with retries on transient failures.
    """
    params = {
        "f": "json",
        "where": where,
        "outFields": out_fields,
        "returnGeometry": "false",
        "orderByFields": "objectid",
        "resultOffset": str(offset),
        "resultRecordCount": str(count),
    }

    last_error: Optional[str] = None

    for attempt in range(3):
        try:
            r = requests.get(ARCGIS_QUERY_URL, params=params, timeout=60)
        except requests.RequestException as e:
            last_error = f"ArcGIS request exception: {e}"
            time.sleep(1.5 * (attempt + 1))
            continue

        # Retry on transient upstream errors
        if r.status_code in (429, 502, 503, 504):
            last_error = f"ArcGIS HTTP {r.status_code}: {r.text[:300]}"
            time.sleep(1.5 * (attempt + 1))
            continue

        if r.status_code != 200:
            raise HTTPException(status_code=502, detail=f"ArcGIS HTTP {r.status_code}: {r.text[:500]}")

        data = r.json()

        # ArcGIS error payload
        if "error" in data:
            msg = (data["error"].get("message") or "")
            details = " ".join(data["error"].get("details") or [])
            combined = (msg + " " + details).strip()

            # Retry generic “Unable to perform query operation” a couple times
            if attempt < 2 and "Unable to perform query operation" in combined:
                last_error = f"ArcGIS error (retryable): {data['error']}"
                time.sleep(1.5 * (attempt + 1))
                continue

            raise HTTPException(status_code=502, detail=f"ArcGIS error: {data['error']}")

        return data

    raise HTTPException(status_code=502, detail=last_error or "ArcGIS query failed after retries")


def _arcgis_fetch_all(where: str, out_fields: str) -> Tuple[List[Dict[str, Any]], str]:
    """
    Fetch all pages for a query.
    """
    all_features: List[Dict[str, Any]] = []
    offset = 0
    page_size = 2000

    while True:
        data = _arcgis_page(where, out_fields, offset, page_size)
        feats = data.get("features", []) or []
        all_features.extend(feats)

        exceeded = bool(data.get("exceededTransferLimit"))
        if not exceeded:
            break
        if len(feats) == 0:
            break
        offset += page_size

    # provenance URL for logging (note: not URL-encoded; ok for notes)
    source_url = f"{ARCGIS_QUERY_URL}?where={where}"
    return all_features, source_url


def _build_where_candidates(city: str, street_number: str, street_keyword: str) -> List[str]:
    """
    Build increasingly permissive where clauses to survive ArcGIS quirks.
    """
    kw = street_keyword.upper().strip()
    num = street_number.strip()

    # 1) Preferred: UPPER + LIKE + city constraint
    w1 = (
        f"true_site_city='{city}' "
        f"AND UPPER(true_site_addr) LIKE '%{num}%' "
        f"AND UPPER(true_site_addr) LIKE '%{kw}%'"
    )

    # 2) Fallback: no UPPER()
    w2 = (
        f"true_site_city='{city}' "
        f"AND true_site_addr LIKE '%{num}%' "
        f"AND true_site_addr LIKE '%{kw}%'"
    )

    # 3) Fallback: relax city constraint; filter city in Python afterwards
    w3 = (
        f"UPPER(true_site_addr) LIKE '%{num}%' "
        f"AND UPPER(true_site_addr) LIKE '%{kw}%'"
    )

    return [w1, w2, w3]


def _arcgis_fetch_all_with_fallback(city: str, street_number: str, street_keyword: str) -> Tuple[List[Dict[str, Any]], str, str]:
    """
    Try multiple where clauses until one succeeds.
    Fail-closed: if none succeed, raise HTTPException.
    """
    last_err: Optional[Exception] = None

    for where in _build_where_candidates(city, street_number, street_keyword):
        try:
            feats, source_url = _arcgis_fetch_all(where=where, out_fields=DEFAULT_OUTFIELDS)

            # If we relaxed city constraint, filter in Python
            if "true_site_city='" not in where:
                feats = [f for f in feats if (f.get("attributes", {}).get("true_site_city") == city)]

            return feats, source_url, where

        except HTTPException as e:
            last_err = e

    raise last_err if last_err else HTTPException(status_code=502, detail="ArcGIS query failed")


def _get_header_map(ws) -> Dict[str, int]:
    headers = [c.value for c in ws[1]]
    return {str(h).strip(): i for i, h in enumerate(headers) if h is not None}


def _cell(ws, row: int, col_name: str, header_map: Dict[str, int]):
    return ws.cell(row=row, column=header_map[col_name] + 1)


@app.get("/health")
def health():
    return {"ok": True}


@app.post("/run/building")
async def run_building(
    workbook: UploadFile = File(...),
    target_name: str = Form(...),
    street_number: str = Form(...),
    street_keyword: str = Form(...),
    city: str = Form("Sunny Isles Beach"),
    state: str = Form("FL"),
    county: str = Form("Miami-Dade"),
    market: str = Form("South Florida"),
    x_api_key: Optional[str] = Header(None, alias="X-API-Key"),
):
    """
    Server-side building run:
    - Fetch parcel unit records (LIKE-first with fallbacks)
    - Append Leads (dedupe on folio+unit)
    - Update Target status (Completed only if records_added >= MIN_TOWER_ROWS_FOR_COMPLETED)
    - Append Run_Log
    - Return updated .xlsx
    """
    _require_api_key(x_api_key)

    # 1) Retrieve parcel features with resilient fallbacks
    features, source_url, where_used = _arcgis_fetch_all_with_fallback(
        city=city, street_number=street_number, street_keyword=street_keyword
    )

    # Fail closed if nothing comes back (don’t produce a “no-op workbook”)
    if not features:
        raise HTTPException(
            status_code=502,
            detail=f"No parcel features returned for target={target_name}. where={where_used}"
        )

    # 2) Load workbook
    content = await workbook.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws_leads = wb["Leads"]
    ws_targets = wb["Targets"]
    ws_log = wb["Run_Log"]

    leads_map = _get_header_map(ws_leads)
    targets_map = _get_header_map(ws_targets)

    # 3) Dedupe set from existing Leads
    existing = set()
    for r in range(2, ws_leads.max_row + 1):
        folio = _cell(ws_leads, r, "folio", leads_map).value
        unit = _cell(ws_leads, r, "unit", leads_map).value
        if folio is None:
            continue
        key = (str(folio).strip(), ("" if unit is None else str(unit).strip()))
        existing.add(key)

    records_added = 0
    duplicates_skipped = 0
    now_dt = datetime.now()
    today = date.today()

    # 4) Append new features into Leads
    for feat in features:
        attr = (feat.get("attributes") or {})
        folio = str(attr.get("folio") or "").strip()
        unit = attr.get("true_site_unit")
        unit_s = "" if unit is None else str(unit).strip()

        if not folio:
            continue

        key = (folio, unit_s)
        if key in existing:
            duplicates_skipped += 1
            continue

        owner_raw = _join_owner(attr.get("true_owner1"), attr.get("true_owner2"), attr.get("true_owner3"))
        owner_type = _classify_owner_type(owner_raw)

        # Build row keyed by your Leads header names
        row = {h: None for h in leads_map.keys()}
        row.update({
            "record_id": str(uuid.uuid4()),
            "created_at": now_dt,
            "last_verified_at": today,
            "status": "New",
            "market": market,
            "county": county,
            "state": state,
            "property_address_1": attr.get("true_site_addr"),
            "unit": None if unit_s == "" else unit_s,
            "city": attr.get("true_site_city"),
            "zip": attr.get("true_site_zip_code"),
            "folio": folio,
            "condo_flag": attr.get("condo_flag"),
            "owner_name_raw": owner_raw,
            "owner_type": owner_type,
            "mailing_address_1": attr.get("true_mailing_addr1"),
            "mailing_address_2": attr.get("true_mailing_addr2"),
            "mailing_address_3": attr.get("true_mailing_addr3"),
            "mailing_city": attr.get("true_mailing_city"),
            "mailing_state": attr.get("true_mailing_state"),
            "mailing_zip": attr.get("true_mailing_zip_code"),
            "mailing_country": attr.get("true_mailing_country"),
            "source_primary": "MDC Open Data",
            "source_urls": source_url,
            "notes": f"parent_folio={attr.get('parent_folio')}" if attr.get("parent_folio") else None,
        })

        out_row = [None] * len(leads_map)
        for col_name, idx in leads_map.items():
            out_row[idx] = row.get(col_name)

        ws_leads.append(out_row)
        existing.add(key)
        records_added += 1

    # Fail closed if we retrieved features but wrote nothing (e.g., all duplicates)
    # In that case, we still return workbook, but mark accordingly.
    # (This is useful on reruns.)
    # You can choose to raise instead, but this is typically fine.

    # 5) Update Targets row
    target_updated = False
    status_value = "Completed" if records_added >= MIN_TOWER_ROWS_FOR_COMPLETED else "Needs Review"

    for r in range(2, ws_targets.max_row + 1):
        val = _cell(ws_targets, r, "value", targets_map).value
        ttype = _cell(ws_targets, r, "target_type", targets_map).value
        if (ttype == "Building") and isinstance(val, str) and (val.strip() == target_name.strip()):
            _cell(ws_targets, r, "status", targets_map).value = status_value
            _cell(ws_targets, r, "completed_at", targets_map).value = today
            notes_cell = _cell(ws_targets, r, "notes", targets_map)
            prev = (notes_cell.value or "")
            extra = f" run={today} added={records_added} skipped={duplicates_skipped} where={where_used}"
            notes_cell.value = (prev + " | " + extra).strip(" |")
            target_updated = True
            break

    # 6) Append Run_Log
    ws_log.append([
        str(uuid.uuid4()),      # run_id
        now_dt,                 # run_datetime
        f"Building: {target_name}",
        records_added,
        duplicates_skipped,
        "" if target_updated else "Target row not found / not updated",
        "Sunny Isles Owner Intel Agent",
        f"status={status_value}; where={where_used}; features_retrieved={len(features)}",
    ])

    # 7) Return updated workbook
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"Owner_Intel_UPDATED_{target_name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\""},
    )
